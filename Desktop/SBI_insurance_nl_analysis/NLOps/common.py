from operator import index
import camelot
import os
import re
import math
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
# from NLOps.next_row import row_start
from next_row import row_start
# from tabula import read_pdf  

def table_extract(file,file_path,base_path):

    p=os.path.join(file_path,file)
    tables = camelot.read_pdf(p,pages='all')
    print(file)
    file=re.sub('(-)|( )','_',file)
    print(file)
    file=file.split('.pdf')[0]
    data_row_start=None
    data_col_start=None
    fl=65

    print(file)

    for i in range(0,len(tables)):

        print(tables[i].parsing_report)
        # df=tables[i].df
        # tables.export(base_path+'/output/raw',file+'.csv')
        
        file_text=''

        if ('nl_4' or 'nl_5' or 'nl_6' or 'nl_7') in file.lower() and 'q1' not in file.lower():

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")
            sheet=workbook['4-7 Top line & Bottom line']

            if 'nl_4_' in file.lower():
                file_name=file+'_table_'+str(i)
                table_frame=tables[i].df
                data4=pd.DataFrame(table_frame)
                for col in list(data4.columns):
                    for row in data4.index:

                        if row==0 and data4[col][row]=='':
                            data[col][row]=data[col-1][row]
                    
                        if row>1 and col>0:
                            data4[col][row]=re.sub('\D','',data4[col][row])

                
            if 'nl_5' in file.lower():
                file_name=file+'_table_'+str(i)
                table_frame=tables[i].df
                data5=pd.DataFrame(table_frame)

                for col in list(data5.columns):
                    for row in data5.index:
                    
                        if row==0 and data5[col][row]=='' and col>1:
                            data5[col][row]=data5[col-1][row]
                    
                        if row>2 and col>0:
                            data5[col][row]=re.sub('\D','',data5[col][row])

            if 'nl_6' in file.lower():
                file_name=file+'_table_'+str(i)

                table_frame=tables[i].df
                data6=pd.DataFrame(table_frame)

                for col in list(data.columns):
                    for row in data6.index:

                        if row==0 and data6[col][row]=='':
                            data6[col][row]=data6[col-1][row]
                    
                        if row>1 and col>0:
                            data6[col][row]=re.sub('\D','',data6[col][row])

            if 'nl_7' in file.lower():
                file_name=file+'_table_'+str(i)

                table_frame=tables[i].df
                data6=pd.DataFrame(table_frame)

                for col in list(data6.columns):
                    for row in data6.index:

                        if col>0:
                            if row==0 and data6[col][row]=='':
                                data[col][row]=data6[col-1][row]

                        if col==0:
                            if data6[col+1][row]=='' and data6[col][row]!='':
                                data6[col+1][row]=data6[col][row]
                                data6[col][row]=''
                    
                        if row>1 and col>1:
                            data6[col][row]=re.sub('\D','',data6[col][row])

                data6=data6.drop(columns=0)
            xl_row_start=row_start(sheet)
            print(xl_row_start)
            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')


        
        if 'nl_7' in file.lower() and 'q1' not in file.lower():
            file_name=file+'_table_'+str(i)

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            cols = list(data.columns)

            for col in list(data.columns):
                for row in data.index:

                    if col>0:
                        if row==0 and data[col][row]=='':
                            data[col][row]=data[col-1][row]

                    if col==0:
                        if data[col+1][row]=='' and data[col][row]!='':
                            data[col+1][row]=data[col][row]
                            data[col][row]=''
                    
                    if row>1 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])

            data=data.drop(columns=0)

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

            sheet=workbook['7 EOM schedule']

            xl_row_start=row_start(sheet)
            print(xl_row_start)

            if 'table_0' in file_name:
                print('hi')
                j=0
                for col in list(data.columns):
                    nl_7_xl_data_row=xl_row_start
                    for row in data.index:
                        
                        if col>=2 and row>=3:
                            if row == 9:
                                #print(row)
                                continue
                            nl_7_xl_data_row=nl_7_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_7_xl_data_row):'T'+str(nl_7_xl_data_row)])
                            # print(row)
                            for x in val_start:
                                for i in x:
                                    # print(i)
                                    if re.search('[B][0-9]',str(i)):
                                       
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)

                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_7_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i.value)
                                            j=j+1
                                        
                                    
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==2:
                                        print(i)
                                        i.value=data[col][row]
                                        print(i.value)
                                    
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==3:
                                        i.value=data[col][row]
                                    if re.search('[H][0-9]{1,2}',str(i)) and col==4:
                                        i.value=data[col][row]
                                    if re.search('[I][0-9]{1,2}',str(i)) and col==5:
                                        i.value=data[col][row]
                                    if re.search('[J][0-9]{1,2}',str(i)) and col==6:
                                        i.value=data[col][row]
                                    
       
            workbook.save(base_path+"/output/"+"final_format.xlsx")

            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')

        if 'nl_27' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)
            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)
            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")
            sheet=workbook['27 Products filed']
            if 'table_0' in file_name :
                print('hi')
                xl_row_start=row_start(sheet)
                print(xl_row_start)

                for col in list(data.columns):
                    nl_27_xl_data_row=xl_row_start
                    print(nl_27_xl_data_row)
                    for row in data.index:
                        
                        if col>=0 and row>=3:
                            nl_27_xl_data_row=nl_27_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_27_xl_data_row):'G'+str(nl_27_xl_data_row)])
                            for x in val_start:
                                for i in x:
                                    # print(i)
                                    if 'B' in str(i):
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)
                                    if 'C' in str(i) and col==0:
                                        # print(data[col][row]
                                        i.value=data[col][row]
                                        print(i.value) 
                                    if 'D' in str(i) and col==1:
                                        i.value=data[col][row]
                                        print(i.value)
                                    if 'E' in str(i) and col==2:
                                        i.value=data[col][row]
                                        # print(i.value)
                                    if 'F' in str(i) and col==4:
                                        i.value=data[col][row]
                                        # print(i.value)
                                    if 'G' in str(i) and col==5:
                                        i.value=data[col][row]
                                        # print(i.value)
            if 'table_1' in file_name:
                print('hi')
                xl_row_start=row_start(sheet)
                print(xl_row_start)

                for col in list(data.columns):
                    nl_27_xl_data_row=xl_row_start
                    print(nl_27_xl_data_row)
                    for row in data.index:
                        
                        if col>=0 and row>=0:
                            nl_27_xl_data_row=nl_27_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_27_xl_data_row):'G'+str(nl_27_xl_data_row)])
                            for x in val_start:
                                for i in x:
                                    # print(i)
                                    if 'B' in str(i):
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)
                                    if 'C' in str(i) and col==0:
                                        # print(data[col][row]
                                        i.value=data[col][row]
                                        print(i.value) 
                                    if 'D' in str(i) and col==1:
                                        i.value=data[col][row]
                                        print(i.value)
                                    if 'E' in str(i) and col==2:
                                        i.value=data[col][row]
                                        # print(i.value)
                                    if 'F' in str(i) and col==4:
                                        i.value=data[col][row]
                                        # print(i.value)
                                    if 'G' in str(i) and col==5:
                                        i.value=data[col][row]
                                        # print(i.value) 

            workbook.save(base_path+"/output/"+"final_format.xlsx")

        if 'nl_33' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            for col in list(data.columns):
                for row in data.index:
                    
                    if row>=3 and col>=2:
                        data[col][row]=re.sub(',','',data[col][row])

            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

            sheet=workbook['33 Reinsurers profile']

            xl_row_start=row_start(sheet)
            print(xl_row_start)

            if 'table_0' in file_name:
                print('hi')
                j=0
                for col in list(data.columns):
                    nl_33_xl_data_row=xl_row_start
                    for row in data.index:
                        
                        if col>=2 and row>=3:
                            if row == 9:
                                #print(row)
                                continue
                            nl_33_xl_data_row=nl_33_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_33_xl_data_row):'J'+str(nl_33_xl_data_row)])
                            # print(row)
                            for x in val_start:
                                for i in x:
                                    # print(i)
                                    if re.search('[B][0-9]',str(i)):
                                       
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)

                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_33_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i.value)
                                            j=j+1
                                        
                                    
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==2:
                                        print(i)
                                        i.value=data[col][row]
                                        print(i.value)
                                    
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==3:
                                        i.value=data[col][row]
                                    if re.search('[H][0-9]{1,2}',str(i)) and col==4:
                                        i.value=data[col][row]
                                    if re.search('[I][0-9]{1,2}',str(i)) and col==5:
                                        i.value=data[col][row]
                                    if re.search('[J][0-9]{1,2}',str(i)) and col==6:
                                        i.value=data[col][row]
                                    
       
            workbook.save(base_path+"/output/"+"final_format.xlsx")

        
        if 'nl_34' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            cols = list(data.columns)

            for col in list(data.columns):
                for row in data.index:
            #         print(row)
                    
                    if col==0:
                        if data[col+1][row]==None and data[col][row]!=None:
                            data[col+1][row]=data[col][row]
                    if col==1:
                        data[col-1][row]=data[col][row].split(' ')[0]
                        # data[col-1][row]=data[col][row].split(' ')[1]
                    
                    if row>1 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])

            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

            sheet=workbook['34 Geo Distribution of Business']

            xl_row_start=row_start(sheet)
            print(xl_row_start)
            if 'hdfc' in file:
                data_row_start=3
                data_col_start=2
            else:
                data_row_start=2
                data_col_start=2



            if 'table_0' in file_name:
                print('hi')
                j=0
                for col in list(data.columns):
                    nl_34_xl_data_row=xl_row_start
                    for row in data.index:


                        
                        if col>=data_col_start and row>=data_row_start:
                            if row in (30,31,32,42,43,44,45,46,47,48) and 'bajaj' in file:
                                continue
                            if row in (31,32,42,43,44,45,46,47,48,49) and 'hdfc' in file:
                                continue
                            nl_34_xl_data_row=nl_34_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_34_xl_data_row):'U'+str(nl_34_xl_data_row)])
                            # print(row)
                            for x in val_start:
                                for i in x:
                                    # print(i)
                                    if re.search('[B][0-9]',str(i)):
                                       
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)
                                    
                                    # if re.search('[C][0-9]{1,2}',str(i)) and col==4:
                                    #     i.value=int(data[col][row])
                                    
                                    # if re.search('[D][0-9]{1,2}',str(i)) and col==4:
                                    #     i.value=data[col][row]

                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_34_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i)
                                            j=j+1
                                    
                                    if 'F' in str(i) and col==2:
                                        # print(data[col][row]
                                        i.value=(data[col][row])
                                        print(i) 
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==4:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'H' in str(i) and col==3:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'I' in str(i) and col==6:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'J' in str(i) and col==7:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'K' in str(i) and col==9:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'L' in str(i) and col==10:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'M' in str(i) and col==11:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'N' in str(i) and col==13:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'O' in str(i) and col==14:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'P' in str(i) and col==15:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'Q' in str(i) and col==16:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'R' in str(i) and col==17:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'S' in str(i) and col==18:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'T' in str(i) and col==19:
                                        i.value=(data[col][row])
                                        # print(i.value)
                                    if 'U' in str(i) and col==21:
                                        i.value=(data[col][row])
                                        # print(i.value)
            workbook.save(base_path+"/output/"+"final_format.xlsx")


        if 'nl_35' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            for col in list(data.columns):
                for row in data.index:
                    if row==0 and data[col][row]=='':
                        data[col][row]=data[col-1][row]
                    
                    if row>1 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])

                        if data[col][row]=='':
                            data[col][row]=data[col][row-1]
                    
                    if row==16 and col==2:
                        data[col][row]=int(data[col][row])+int(data[col][row-1])

                    if row==16 and col==3:
                        data[col][row]=int(data[col][row])+int(data[col][row-1])
                    

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

            # sheet=workbook['35 Business returns across LOBs']
            sheet=workbook['35 Business returns across LOBs']

            xl_row_start=row_start(sheet)
            print(xl_row_start)

            if 'table_0' in file_name:
                print('hi')
                TOTAL_A=0
                TOTAL_B=0
                j=0
                f=open(base_path+'/NLOps/row/nl_35_row.json')
                d=list(json.load(f).values())
                for col in list(data.columns):
                    nl_35_xl_data_row=xl_row_start
                    # print(nl_35_xl_data_row)
                    for row in data.index:
                        # if row==15:
                        #     continue
                        
                        if col>=2 and col <=3 and row>=1:
                            nl_35_xl_data_row=nl_35_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_35_xl_data_row):'F'+str(nl_35_xl_data_row)])
                            for x in val_start:
                                for i in x:

                                    if re.search('[B][0-9]',str(i)):                                       
                                        i.value=file_name.split('_')[0]
                                        
                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i)
                                            print(j)
                                            j=j+1
                                    if row==16:
                                        continue
                                    
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        print(i)
                                        i.value=data[col][row+1]
                                        TOTAL_A=TOTAL_A+int(data[col][row+1])

                                        if 'Total' in d:
                                            sheet['E'+str(nl_35_xl_data_row+1)].value=TOTAL_A
                                    
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==3:
                                        i.value=(data[col][row+1])
                                        TOTAL_B=TOTAL_B+int(data[col][row+1])
                                        if 'Total' in d:
                                            sheet['F'+str(nl_35_xl_data_row+1)].value=TOTAL_B
                                            sheet['B'+str(nl_35_xl_data_row+1)].value=file_name.split('_')[0]  
                    
            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')
            workbook.save(base_path+"/output/"+"final_format.xlsx")
          



        if 'nl_366' in file.lower() and 'q2' not in file.lower():
            file_name=file+'_table_'+str(i)

            print()

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            cols = list(data.columns)
            # print(cols)
            # print(data[8][2])
            data_row_start=data[0][8]
            for col in list(data.columns):
                for row in data.index:
            #         print(row)

                    if col>0:
                        if row==1 and data[col][row]=='':
                            data[col][row]=data[col-1][row]

                    if col==0 and row<18 and row>5:
                        
                        if data[col][row+1]=='' and data[col][row]!='':
                            data[col][row+1]=data_row_start+chr(fl)
                            fl=fl+1
                            
                    
                    if row>2 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])
                    
                    if col==8 and 'q2' in file.lower():
                        if row==2:
                            s=data[col][row].split('Premium')
                            data[col][row]=s[0]
                            data[col+1][row]='Premium '+s[1]
                        if row>2:
                            data[col+1][row]=data[col][row]



            data=data.drop([0])
            data_row_start=None
            # data.to_excel(base_path+'/output/formatted/'+file_name+'_new'+'.xlsx')

          
           
        if 'nl_37' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)
            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)
            # data=data.append(data[])
            if 'hdfc' in file:
                data.loc[8.5] = ['','','','','','','','','','','','','','','','','','','','','']
                data = data.sort_index().reset_index(drop=True)

            cols = list(data.columns)

            for col in list(data.columns):
                for row in data.index:
            #         print(row)
                    
                    if row>0 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])

            workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

            sheet=workbook['37.A. Number of Claims']
            # print(sheet['C3'].value)

            if 'table_0' in file_name:
                workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

                sheet=workbook['37.A. Number of Claims']
                xl_row_start=row_start(sheet)
                print(xl_row_start)
                j=0
                for col in list(data.columns):
                    nl_37_xl_data_row=xl_row_start
                    # print(nl_35_xl_data_row)
                    for row in data.index:
                        
                        if col>=2 and row>=1:
                            nl_37_xl_data_row=nl_37_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_37_xl_data_row):'T'+str(nl_37_xl_data_row)])
                            for x in val_start:
                                for i in x:

                                    if re.search('[B][0-9]',str(i)):   
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)

                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_37_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i.value)
                                            j=j+1

                                    

                                    if re.search('[C][0-9]{1,2}',str(i)) and col==0:
                                        i.value=data[col][row]
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        i.value=data[col][row]
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==3:
                                        i.value=(data[col][row])
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        i.value=data[col][row]                                    
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==4:
                                        i.value=(data[col][row])
                                    if re.search('[H][0-9]{1,2}',str(i)) and col==6:
                                        i.value=data[col][row]
                                    if re.search('[I][0-9]{1,2}',str(i)) and col==7:
                                        i.value=(data[col][row])
                                    if re.search('[J][0-9]{1,2}',str(i)) and col==9:
                                        i.value=data[col][row]                                    
                                    if re.search('[K][0-9]{1,2}',str(i)) and col==10:
                                        i.value=(data[col][row])
                                    if re.search('[L][0-9]{1,2}',str(i)) and col==11:
                                        i.value=(data[col][row])
                                    if re.search('[M][0-9]{1,2}',str(i)) and col==13:
                                        i.value=(data[col][row])
                                    if re.search('[N][0-9]{1,2}',str(i)) and col==14:
                                        i.value=(data[col][row])
                                    if re.search('[O][0-9]{1,2}',str(i)) and col==15:
                                        i.value=(data[col][row])
                                    if re.search('[P][0-9]{1,2}',str(i)) and col==16:
                                        i.value=(data[col][row])
                                    if re.search('[Q][0-9]{1,2}',str(i)) and col==17:
                                        i.value=(data[col][row])
                                    if re.search('[R][0-9]{1,2}',str(i)) and col==18:
                                        i.value=(data[col][row])
                                    if re.search('[S][0-9]{1,2}',str(i)) and col==19:
                                        i.value=(data[col][row])
                                    if re.search('[T][0-9]{1,2}',str(i)) and col==20:
                                        i.value=(data[col][row])
            
            workbook.save(base_path+"/output/"+"final_format.xlsx")
            
            if 'table_1' in file_name:
                workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

                sheet=workbook['37.B. Amount of claims']
                xl_row_start=row_start(sheet)
                print(xl_row_start)
                j=0
                for col in list(data.columns):
                    nl_37_xl_data_row=xl_row_start
                    # print(nl_35_xl_data_row)
                    for row in data.index:
                        nl_37_xl_data_row=nl_37_xl_data_row+1
                        if col>=2 and row>=1:
                            val_start=list(sheet['B'+str(nl_37_xl_data_row):'T'+str(nl_37_xl_data_row)])
                            for x in val_start:
                                for i in x:

                                    if re.search('[B][0-9]',str(i)):   
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)
                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_37_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i.value)
                                            j=j+1
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        i.value=data[col][row]
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==3:
                                        i.value=(data[col][row])
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        i.value=data[col][row]                                    
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==4:
                                        i.value=(data[col][row])
                                    if re.search('[H][0-9]{1,2}',str(i)) and col==6:
                                        i.value=data[col][row]
                                    if re.search('[I][0-9]{1,2}',str(i)) and col==7:
                                        i.value=(data[col][row])
                                    if re.search('[J][0-9]{1,2}',str(i)) and col==9:
                                        i.value=data[col][row]                                    
                                    if re.search('[K][0-9]{1,2}',str(i)) and col==10:
                                        i.value=(data[col][row])
                                    if re.search('[L][0-9]{1,2}',str(i)) and col==11:
                                        i.value=(data[col][row])
                                    if re.search('[M][0-9]{1,2}',str(i)) and col==13:
                                        i.value=(data[col][row])
                                    if re.search('[N][0-9]{1,2}',str(i)) and col==14:
                                        i.value=(data[col][row])
                                    if re.search('[O][0-9]{1,2}',str(i)) and col==15:
                                        i.value=(data[col][row])
                                    if re.search('[P][0-9]{1,2}',str(i)) and col==16:
                                        i.value=(data[col][row])
                                    if re.search('[Q][0-9]{1,2}',str(i)) and col==17:
                                        i.value=(data[col][row])
                                    if re.search('[R][0-9]{1,2}',str(i)) and col==18:
                                        i.value=(data[col][row])
                                    if re.search('[S][0-9]{1,2}',str(i)) and col==19:
                                        i.value=(data[col][row])
                                    if re.search('[T][0-9]{1,2}',str(i)) and col==20:
                                        i.value=(data[col][row])       

                workbook.save(base_path+"/output/"+"final_format.xlsx")
                # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')

        if 'nl_39' in file.lower() and 'q2' in file.lower():
            file_name=file+'_table_'+str(i)

            table_frame=tables[i].df
            data=pd.DataFrame(table_frame)

            for col in list(data.columns):
                for row in data.index:
            #         print(row)
                    
                    if row>2 and col>1:
                        data[col][row]=re.sub('\D','',data[col][row])
                    
                    if row==0 and data[col][row]=='':
                        data[col][row]=data[col-1][row]

            data=data.drop([0])


            if 'table_0' in file_name:
                workbook = load_workbook(base_path+"/output/"+"final_format.xlsx")

                sheet=workbook['39 Ageing of claims']
                xl_row_start=row_start(sheet)
                print(xl_row_start)
                j=0
                for col in list(data.columns):
                    nl_39_xl_data_row=xl_row_start
                    # print(nl_35_xl_data_row)
                    for row in data.index:
                        
                        if col>=2 and row>=3:
                            nl_39_xl_data_row=nl_39_xl_data_row+1
                            val_start=list(sheet['B'+str(nl_39_xl_data_row):'T'+str(nl_39_xl_data_row)])
                            for x in val_start:
                                for i in x:

                                    if re.search('[B][0-9]',str(i)):   
                                        i.value=file_name.split('_')[0]
                                        # print(i.value)
                                    if re.search('[D][0-9]{1,2}',str(i)):
                                        # print(i)
                                        f=open(base_path+'/NLOps/row/nl_39_row.json')
                                        d=list(json.load(f).values())
                                        if j<len(d):
                                            i.value=d[j]
                                            print(i.value)
                                            j=j+1
                                    if re.search('[E][0-9]{1,2}',str(i)) and col==2:
                                        i.value=data[col][row]
                                    if re.search('[F][0-9]{1,2}',str(i)) and col==3:
                                        i.value=(data[col][row])                                    
                                    if re.search('[G][0-9]{1,2}',str(i)) and col==4:
                                        i.value=(data[col][row])
                                    if re.search('[H][0-9]{1,2}',str(i)) and col==5:
                                        i.value=data[col][row]
                                    if re.search('[I][0-9]{1,2}',str(i)) and col==6:
                                        i.value=(data[col][row])
                                    if re.search('[J][0-9]{1,2}',str(i)) and col==7:
                                        i.value=data[col][row]                                    
                                    if re.search('[K][0-9]{1,2}',str(i)) and col==8:
                                        i.value=(data[col][row])
                                    if re.search('[L][0-9]{1,2}',str(i)) and col==9:
                                        i.value=(data[col][row])
                                    if re.search('[M][0-9]{1,2}',str(i)) and col==10:
                                        i.value=(data[col][row])
                                    if re.search('[N][0-9]{1,2}',str(i)) and col==11:
                                        i.value=(data[col][row])
                                    if re.search('[O][0-9]{1,2}',str(i)) and col==12:
                                        i.value=(data[col][row])
                                    if re.search('[P][0-9]{1,2}',str(i)) and col==13:
                                        i.value=(data[col][row])
                                    if re.search('[Q][0-9]{1,2}',str(i)) and col==14:
                                        i.value=(data[col][row])
                                    if re.search('[R][0-9]{1,2}',str(i)) and col==15:
                                        i.value=(data[col][row])
                                    if re.search('[S][0-9]{1,2}',str(i)) and col==16:
                                        i.value=(data[col][row])
                                    if re.search('[T][0-9]{1,2}',str(i)) and col==17:
                                        i.value=(data[col][row])
            
            workbook.save(base_path+"/output/"+"final_format.xlsx")
            # data.to_excel(base_path+'/output/formatted/'+file_name+'.xlsx')
            

        if 'nl_47' in file.lower():
            file_name=file+'_table_'+str(i)
          
            # tables[i].to_excel(base_path+'/output/raw/'+file_name+'.xlsx')


def table_format(file,file_path,base_path):
    path=file_path+'/'+file
    wb=load_workbook(path)
    sheet=wb.active
    print(sheet['C4'].value)

# def combine_nl4_nl5_nl6_nl7():

f=open('config.json')
base_path=json.load(f)['base_path']
# savepath=base_path+'/output/raw'
file='hdfc_21_22_q2_NL-37-CLAIMS DATA .pdf'
# # table_extract(file,savepath,base_path)
# table_format(file,savepath,base_path)
file_path=base_path+'/crawler/hdfc/21_22'
table_extract(file,file_path,base_path)